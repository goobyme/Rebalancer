import requests
import bs4
import os
import re
import openpyxl
import shutil
import datetime
import time
import pprint
import alpha_vantage

print('Verifying Directory...\n')
os.makedirs('Fund Data', exist_ok=True)
os.chdir('Fund Data')


def quote_scraper(ticker):

    # Input; Ticker Input: return price of stock  TODO plus more info
    # Note: function is embedded into Update Portfolio Data and not part of Portfolio Main class
    base_url = 'https://finance.google.com/finance?q=' + ticker  # TODO insert try statement for error handling?
    print('Downloading Quote for %s...' % ticker)
    res = requests.get(base_url)
    res.raise_for_status()

    # Regex used as bs4 has difficulty picking up only the ticker price text
    price_check = re.compile(r'(?<=>)\d+\.\d+')

    gfinance_source = bs4.BeautifulSoup(res.text, 'lxml')
    raw_price = gfinance_source.select('span[class=pr]')

    # TODO Insert more data scrapping (asset classes could be easy for stocks (market cap) think about funds too)

    # Error handling for anomalies in GFinance HTML
    try:
        ticker_price = price_check.search(str(raw_price[0]))
    except IndexError:
        print('Could not find quote data')
        return None

    return ticker_price.group(0)


class PortfolioMain:

    def __init__(self, profile):
        # On initial run creates portfolio, or can take pre-existing excel table

        self.quote_dic = {}             # Note these object var only for those that are reset each use
        self.to_email_list = []
        self.ideal_alloc = {}       # TODO FIll with actual data
        self.raw_message_list = []
        self.profile = profile.replace(' ', '').lower()   # Proper formatting of filename
        self.file_name = '%sfund.xlsx' % profile

        if not os.path.isfile(self.file_name):
            shutil.copy('sampledonoteditfund.xlsx', self.file_name)
            print('Fill workbook with ticker data, all other columns are optional.')
            # Long-term look for ways to import data from other ways to lessen need for data entry
            pass

        # recall ticker and share value from variable storage file, TODO Make reading spread conditional

    def spread2binaries(self, filepath):
        # Will read and convert ticker share count data to binaries that can be saved to file and read directly in py
        # might not be needed if can save entire object
        pass

    def update_portfolio_data(self):

        # Input; spreadsheet or raw stock info. Output stock quotes written in spreadsheet and self.quote_dic
        fundspread = openpyxl.load_workbook(self.file_name)
        type('wb')
        sheet = fundspread.get_sheet_by_name('Sheet1')  # TODO Increase flexibility of sheet selector

        # Parses ticker values from spreadsheet or raw info and runs QuoteScraper function to recieve quote info
        # TODO add functionality to read saved stock binaries instead of wb
        for i in range(2, sheet.max_row + 1):
            ticker = sheet.cell(row=i, column=1).value
            stock_quote = quote_scraper(ticker)
            sheet['F' + str(i)] = float(stock_quote)
            self.quote_dic[ticker] = float(stock_quote)

        # Denotes time of update for record keeping purposes ! Can be expanded upon long term
        update_time = datetime.datetime.fromtimestamp(time.time())
        sheet['M1'] = update_time.strftime('%Y/%m/%d %H:%M')

        # Saves and closes spreadsheet.
        try:
            fundspread.save(self.file_name)
        except PermissionError:
            print('Please close excel-sheet')
            pass
        fundspread.close()

    def rebalancer(self):

        # Input, spreadsheet or raw stock and share info. Output raw message data in self.raw_message_list
        # Opens workbook and creates necessary dictionaries (replace with array when possible)
        fundspread = openpyxl.load_workbook(self.file_name)
        type('wb')
        sheet = fundspread.get_sheet_by_name('Sheet1')      # TODO Find way to reduce redundancy across methods

        notation_dic = {}       # Keys = Ticker Values = Notational value of each position
        actual_alloc_dic = {}   # Keys = Ticker Values = Actual allocation in percentage (0.01 = 1%)
        ideal_alloc = {}        # import from alloc setter? TODO add ideal alloc in init
        differentials = {}      # Keys = Ticker Values = Difference between allocations in percent of total portfolio
        accumulate_diff = []    # List of all diff above threshold, used to calc sum of rebalancing trades

        # Fills information in above table through parsing excel sheet and calculating relevant values
        for i in range(2, sheet.max_row + 1):
            ticker = str(sheet.cell(row=i, column=1).value)
            share_count = float(sheet.cell(row=i, column=5).value)      # TODO Build to replace with saved database

            notation_dic[ticker] = round(share_count * self.quote_dic[ticker], 2)

        pf_value = sum(notation_dic.values())   # Calculated total portfolio value
        for ticker in notation_dic.keys():
            actual_alloc_dic[ticker] = notation_dic[ticker]/pf_value
            differentials[ticker] = actual_alloc_dic[ticker] - self.ideal_alloc[ticker]

        # Find all objects in diff list greater than threshold and appends raw ouput string and accumulated diff
        def iterating_rebalance(deltalow, deltahigh):

            for iter_tick in differentials.keys():
                if differentials[iter_tick] in range(deltalow, deltahigh):
                    share2sell = (notation_dic[iter_tick] - (ideal_alloc[iter_tick] * pf_value)) / self.quote_dic[iter_tick]
                    output = ["SELL", iter_tick, differentials[iter_tick], share2sell]

                    accumulate_diff.append(differentials[iter_tick])
                    self.raw_message_list.append(output)

                elif differentials[iter_tick] in range(-deltahigh, - deltalow):
                    share2sell = ((ideal_alloc[iter_tick] * pf_value) - notation_dic[iter_tick]) / self.quote_dic[iter_tick]
                    output = ["BUY", iter_tick, differentials[iter_tick], share2sell]

                    accumulate_diff.append(differentials[iter_tick])
                    self.raw_message_list.append(output)

        # Initial loop of function for basic .05 threshold
        h = 1
        l = .05
        iterating_rebalance(l, h)

        # Continual loops with smaller and smaller thresholds (without repeats) to get accumulated diff lower
        while abs(sum(accumulate_diff)) > 0.025:
            h = l
            l /= 2
            iterating_rebalance(l, h)

    def mail_machine(self, email):

        if not self.raw_message_list:
            print('No messages to send')
            return None

        # Construct nicely formated messages from raw output lists
        for output in self.raw_message_list:
            assert output[0] == "BUY" or "SELL"
            if output[0] == "SELL":
                message = """%s is %s percent above the target allocation with a notational value of %s dollars.  
                           To reach to target notational value of %s dollars, sell %s shares of %s."""
                self.to_email_list.append(message)

            else:
                message = """%s is %s percent below the target allocation with a notational value of %s dollars.  
                            To reach to target notational value of %s dollars, buy %s shares of %s."""
                self.to_email_list.append(message)

    def change_alloc(self):

        for ticker in self.ideal_alloc.keys():
            while True:
                print('Current Allocation: %s. What is your new allocation?' % (self.ideal_alloc[ticker]*100))
                new_alloc = input()
                if not new_alloc:
                    break
                elif new_alloc in range(0, 1):
                    self.ideal_alloc[ticker] = new_alloc
                    break
                else:
                    print('Invalid input. Allocation must be between 0 and 1')
                    continue

        while True:
            print("Input new ticker value (press ENTER when done):")
            new_ticker = input()
            if not new_ticker:
                break
            print('Input allocation for new ticker:')
            new_alloc = input()
            self.ideal_alloc[new_ticker] = new_alloc

        print('\n')
        pprint.pprint(self.ideal_alloc)

    def display_alloc(self):

        pprint.pprint(self.ideal_alloc)
        pass

    def save2file(self):

        pass

# Include sleep function here or in global function

# TODO Build outer code structure. Program needs to open itself every week and send email if needed then sleep again
# TODO Build inherited classes for alternate ideal allocations, or some other means of introducing differing alloc
